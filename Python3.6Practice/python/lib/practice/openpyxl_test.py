#!/usr/bin/python
# coding: utf-8
'''
@description: 
@date: Created on 2017年4月21日
@author: Mr.LiuDC
@email: 1911939348@qq.com
'''
from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
# import datetime
# ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")